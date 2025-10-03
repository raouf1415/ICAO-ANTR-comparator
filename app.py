import io
import re
from typing import List, Dict, Any

import streamlit as st
import pandas as pd
import numpy as np

# Optional readers
try:
    import pdfplumber
except Exception:
    pdfplumber = None

try:
    import docx
except Exception:
    docx = None

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.neighbors import NearestNeighbors

# --------- Utilities ---------
CLAUSE_SPLIT_PATTERN = re.compile(
    r"(?:^\s*(?:Annex|Appendix|Chapter|Section|Subpart|Part)\s+\d+[A-Za-z]*\.?|\n\s*\d+(?:\.\d+){0,4}\s+)",
    re.IGNORECASE | re.MULTILINE
)

def read_any(uploaded) -> str:
    name = uploaded.name.lower()
    raw = uploaded.read()
    uploaded.seek(0)
    if name.endswith(".txt"):
        return raw.decode("utf-8", errors="ignore")
    if name.endswith(".pdf"):
        if pdfplumber is None:
            raise RuntimeError("PDF support missing. Please ensure pdfplumber is installed.")
        with pdfplumber.open(io.BytesIO(raw)) as pdf:
            texts = [(p.extract_text() or "") for p in pdf.pages]
        return "\n".join(texts)
    if name.endswith(".docx"):
        if docx is None:
            raise RuntimeError("DOCX support missing. Please ensure python-docx is installed.")
        d = docx.Document(io.BytesIO(raw))
        return "\n".join(p.text for p in d.paragraphs)
    # Fallback
    return raw.decode("utf-8", errors="ignore")

def normalize_ws(s: str) -> str:
    """Normalize whitespace while preserving line breaks for clause splitting.

    - Convert all line endings to "\n"
    - Collapse horizontal whitespace (spaces, tabs) but keep newlines
    - Trim each line and collapse excessive blank lines
    """
    s = (s or "")
    # Normalize line endings
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    # Collapse horizontal whitespace but preserve newlines
    s = re.sub(r"[^\S\n]+", " ", s)
    # Trim each line
    s = "\n".join(line.strip() for line in s.split("\n"))
    # Collapse multiple blank lines
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def split_clauses(text: str) -> List[Dict[str, Any]]:
    text = normalize_ws(text)
    if not text:
        return []
    positions = [m.start() for m in CLAUSE_SPLIT_PATTERN.finditer("\n" + text)]
    positions = sorted(set([0] + positions + [len(text)]))
    chunks = []
    for i in range(len(positions) - 1):
        chunk = text[positions[i]:positions[i+1]].strip()
        if not chunk:
            continue
        m = re.match(r"^(Annex|Appendix|Chapter|Section|Subpart|Part)\s+([0-9A-Za-z\.]+)", chunk, re.I)
        n = re.match(r"^(\d+(?:\.\d+){0,4})", chunk)
        if m:
            cid = f"{m.group(1).title()} {m.group(2)}"
        elif n:
            cid = n.group(1)
        else:
            cid = " ".join(chunk.split()[:6]) + ("..." if len(chunk.split())>6 else "")
        chunks.append({"id": cid, "text": chunk})
    counts = {}
    for c in chunks:
        b = c["id"]
        counts[b] = counts.get(b, 0) + 1
        if counts[b] > 1:
            c["id"] = f"{b} ({counts[b]})"
    return chunks

def align(src, tgt, min_sim=0.35, top_k=1) -> pd.DataFrame:
    """Align clauses using TF-IDF + top-k cosine neighbors without building full matrix.

    This avoids O(N*M) memory/time of a full similarity matrix and scales better.
    """
    A_texts = [c["text"] for c in src]
    B_texts = [c["text"] for c in tgt]
    if not A_texts or not B_texts:
        return pd.DataFrame(columns=[
            "Source_ID",
            "Source_Text",
            "Matched_Target_ID",
            "Matched_Target_Text",
            "Similarity",
            "Gap?",
        ])

    vec = TfidfVectorizer(ngram_range=(1, 2))
    X = vec.fit_transform(A_texts + B_texts)
    A = X[: len(A_texts)]
    B = X[len(A_texts) :]

    # Use cosine distance (1 - cosine similarity). Supports sparse input with brute algorithm.
    n_neighbors = min(max(1, top_k), B.shape[0])
    nn = NearestNeighbors(n_neighbors=n_neighbors, metric="cosine", algorithm="brute")
    nn.fit(B)
    distances, indices = nn.kneighbors(A, return_distance=True)

    rows = []
    for i, sc in enumerate(src):
        neighbor_idxs = indices[i]
        neighbor_dists = distances[i]
        added = False
        for k, j in enumerate(neighbor_idxs):
            sim = float(1.0 - neighbor_dists[k])
            if sim >= min_sim:
                rows.append(
                    {
                        "Source_ID": sc["id"],
                        "Source_Text": sc["text"],
                        "Matched_Target_ID": tgt[int(j)]["id"],
                        "Matched_Target_Text": tgt[int(j)]["text"],
                        "Similarity": round(sim, 4),
                        "Gap?": "Aligned",
                        "_Source_Order": i,
                    }
                )
                added = True
        if not added:
            j0 = int(neighbor_idxs[0])
            sim0 = float(1.0 - neighbor_dists[0])
            rows.append(
                {
                    "Source_ID": sc["id"],
                    "Source_Text": sc["text"],
                    "Matched_Target_ID": tgt[j0]["id"],
                    "Matched_Target_Text": tgt[j0]["text"],
                    "Similarity": round(sim0, 4),
                    "Gap?": "Potential Gap",
                    "_Source_Order": i,
                }
            )

    df = pd.DataFrame(rows)
    if not df.empty:
        df = (
            df.sort_values(["_Source_Order", "Similarity"], ascending=[True, False])
            .drop(columns=["_Source_Order"])  # keep original source order
            .reset_index(drop=True)
        )
    return df

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    import io
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill
    out = io.BytesIO()
    # Ensure index is simple integer to avoid Excel writer issues on strange indexes
    df_to_write = df.reset_index(drop=True)
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df_to_write.to_excel(w, index=False, sheet_name="Matrix")
        ws = w.sheets["Matrix"]
        for i, _ in enumerate(df_to_write.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 24
        # Use ARGB 8-digit color code for compatibility with openpyxl
        gap_fill = PatternFill(start_color="FFFFF3CD", end_color="FFFFF3CD", fill_type="solid")
        gap_idx = df_to_write.columns.get_loc("Gap?") if "Gap?" in df_to_write.columns else None
        for r in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if gap_idx is not None and r[gap_idx].value == "Potential Gap":
                for cell in r:
                    cell.fill = gap_fill
    return out.getvalue()

# --------- UI ---------
st.set_page_config(page_title="ICAO / ANTR Comparator", layout="wide")
st.title("📄 ICAO / ANTR Document Comparator")

st.sidebar.header("Settings")
min_sim = st.sidebar.slider("Minimum Similarity (Gap Threshold)", 0.0, 1.0, 0.35, 0.01)
top_k = st.sidebar.slider("Top matches per clause", 1, 3, 1)
split_hint = st.sidebar.text_input("Custom regex for clause split (optional)", "")

c1, c2 = st.columns(2)
with c1:
    src_file = st.file_uploader("Primary Document (e.g., ANTR)", type=["pdf","docx","txt"], key="src")
with c2:
    tgt_file = st.file_uploader("Reference Document (e.g., ICAO Annex)", type=["pdf","docx","txt"], key="tgt")

if src_file and tgt_file:
    with st.spinner("Reading & parsing documents…"):
        try:
            src_text = read_any(src_file)
            tgt_text = read_any(tgt_file)
        except Exception as e:
            st.error(f"File read error: {e}")
            st.stop()

        if split_hint.strip():
            try:
                # Anchor custom hint to start-of-line by default to avoid catastrophic backtracking
                hint = split_hint
                if not hint.startswith("^"):
                    hint = r"^" + hint
                CLAUSE_SPLIT_PATTERN = re.compile(hint, re.IGNORECASE | re.MULTILINE)
            except re.error as ex:
                st.warning(f"Ignoring invalid regex: {ex}")

        src = split_clauses(src_text)
        tgt = split_clauses(tgt_text)

        st.success(f"Parsed {len(src)} clauses (Primary) and {len(tgt)} clauses (Reference).")

    with st.spinner("Computing alignments…"):
        df = align(src, tgt, min_sim=min_sim, top_k=top_k)

    st.subheader("Alignment / Gap Matrix")
    st.dataframe(df, use_container_width=True, height=520)

    st.download_button("⬇️ Download Excel", data=to_excel_bytes(df), file_name="gap_matrix.xlsx")
    st.download_button("⬇️ Download CSV", data=df.to_csv(index=False).encode("utf-8"), file_name="gap_matrix.csv")

else:
    st.info("Upload two documents (PDF, DOCX, or TXT) to begin.")